import os, time, random, zipfile, requests, re, traceback
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException

# --- настройки цикла подтверждения ---
CONFIRM_ATTEMPTS = 6        # сколько раз ждём по 90 секунд
CONFIRM_WAIT = 90           # длительность одного ожидания, секунд
AUTH_PAGE_TIMEOUT = 60      # сколько ждём переход на mobile-id-auth
RESERVE_TIMEOUT = 30        # сколько ждём страницу с номером после Элемента 6
CLOSE_FAILED_WINDOW = True  # закрывать окно, если оно не стало успешным

REPEAT_BTN_XPATH = '//*[@id="root"]/div/div/div/main/div/div/div/div[2]/button'

def pause():
    time.sleep(random.uniform(1, 2))

def make_proxy_auth_extension(host, port, user, pwd, path):
    manifest = """{
        "version":"1.0.0","manifest_version":2,"name":"Proxy Auth",
        "permissions":["proxy","tabs","unlimitedStorage","storage","<all_urls>","webRequest","webRequestBlocking"],
        "background":{"scripts":["background.js"]},"minimum_chrome_version":"76.0.0"}"""
    background = f"""
    var config={{mode:"fixed_servers",rules:{{singleProxy:{{scheme:"http",host:"{host}",port:parseInt({port})}},bypassList:["localhost","127.0.0.1"]}}}};
    chrome.proxy.settings.set({{value:config,scope:"regular"}},function(){{}});
    function cb(d){{return{{authCredentials:{{username:"{user}",password:"{pwd}"}}}}}};
    chrome.webRequest.onAuthRequired.addListener(cb,{{urls:["<all_urls>"]}},["blocking"]);
    """
    with zipfile.ZipFile(path,"w") as zp:
        zp.writestr("manifest.json",manifest)
        zp.writestr("background.js",background)

def is_ready(el):
    """Кнопка реально доступна для клика: видима, включена, без disabled-признаков."""
    try:
        disabled_attr = el.get_attribute('disabled')
        aria_disabled = el.get_attribute('aria-disabled')
        btn_class = (el.get_attribute('class') or '').lower()
        return (
            el.is_displayed()
            and el.is_enabled()
            and disabled_attr is None
            and aria_disabled not in ('true', 'True')
            and 'disabled' not in btn_class
        )
    except WebDriverException:
        # элемент пропал со страницы прямо во время проверки — считаем недоступным
        return False

def click_multi(driver, selectors, name, timeout=4):
    end = time.time() + timeout
    while time.time() < end:
        for sel in selectors:
            try:
                el = driver.find_element(By.XPATH, sel)
                if el.is_enabled() and el.is_displayed():
                    el.click()
                    print(f"Клик по {name}: {sel}")
                    pause()
                    return True
            except:
                continue
        time.sleep(0.2)
    print(f"{name} не найден")
    return False

def wait_confirmation(driver):
    """
    CONFIRM_ATTEMPTS раз: ждём CONFIRM_WAIT секунд и жмём 'повторить ещё'.
    Если после очередного ожидания кнопки нет или она недоступна — окно успешное.
    Возвращает True, если окно успешное.
    """
    for attempt in range(1, CONFIRM_ATTEMPTS + 1):
        print(f"Ожидание {attempt}/{CONFIRM_ATTEMPTS} — {CONFIRM_WAIT} секунд...")
        time.sleep(CONFIRM_WAIT)

        btns = driver.find_elements(By.XPATH, REPEAT_BTN_XPATH)
        if not btns:
            print(f"Кнопка отсутствует (попытка {attempt}) — окно успешное")
            return True

        btn = btns[0]
        if not is_ready(btn):
            print(f"Кнопка есть, но недоступна (попытка {attempt}) — окно успешное")
            return True

        try:
            btn.click()
            print(f"Нажал 'повторить ещё' (попытка {attempt})")
        except WebDriverException as e:
            # кнопка исчезла/перекрыта в момент клика — значит доступной она не осталась
            print(f"Клик по 'повторить ещё' не прошёл (попытка {attempt}): {type(e).__name__} — окно успешное")
            return True

    print(f"Все {CONFIRM_ATTEMPTS} попыток кнопка оставалась доступной — окно НЕуспешное")
    return False

base_dir = os.path.dirname(os.path.abspath(__file__))
excel_file=os.path.join(base_dir,"clients.xlsx")
log_file=os.path.join(base_dir,"все логи.txt")

wb=load_workbook(excel_file); ws=wb.active
rows=list(ws.iter_rows(min_row=2,values_only=True))

proxy_input=input("Введите прокси (IP:PORT:LOGIN:PASS[IP_CHANGE_URL]): ").strip()
if "[" in proxy_input:
    base,ip_change_url=proxy_input.split("[",1)
    ip_change_url=ip_change_url.strip("]")
else:
    base=proxy_input; ip_change_url=None
ip,port,login,password=base.split(":")
ext_path=os.path.join(base_dir,"proxy_auth_ext.zip")
make_proxy_auth_extension(ip,port,login,password,ext_path)

row_index=0
while row_index<len(rows):
    if ip_change_url:
        try: requests.get(ip_change_url,timeout=10)
        except: pass
    edge_options=EdgeOptions()
    edge_options.add_extension(ext_path)
    edge_options.page_load_strategy = "eager"
    driver=webdriver.Edge(options=edge_options)
    wait=WebDriverWait(driver,15)

    num = pas = current_url = digits = None
    window_ok = False
    reached_confirmation = False  # дошли ли до цикла подтверждения 6x90

    try:
        driver.get("https://saratov.beeline.ru/basket/")
        driver.set_window_size(1920,1080)
        time.sleep(3)

        # Элемент 1
        click_multi(driver, [
            '//*[@id="esim"]',
            '//*[@id="root"]/div/div/div[2]/div/div/section/label[2]/span[2]',
            '//*[@id="root"]/div/div/div[2]/div/div/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div/div[2]/div'
        ], "Элемент 1")

        # Элемент 2
        click_multi(driver, [
            '//*[@id="root"]/div/div/div[2]/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/button[1]/span'
        ], "Элемент 2")

        # Элемент 3 (все твои XPATH‑ы)
        click_multi(driver, [
            '/html/body/div[17]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p',
            '/html/body/div[17]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button',
            '/html/body/div[18]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p',
            '/html/body/div[16]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p',
            '/html/body/div[14]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p',
            '/html/body/div[19]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button',
            '/html/body/div[6]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p',
            '/html/body/div[3]/div[2]/div/div/div[2]/div/div[11]/div/div/div[2]/div[2]/button/p'
        ], "Элемент 3")

        # Элемент 4 (все твои XPATH‑ы)
        click_multi(driver, [
            '/html/body/div[18]/div[2]/div/div/div/div/div/div[3]/button/p',
            '/html/body/div[17]/div[2]/div/div/div/div/div/div[3]/button/p',
            '/html/body/div[19]/div[2]/div/div/div/div/div/div[3]/button/p',
            '/html/body/div[6]/div[2]/div/div/div/div/div/div[3]/button/p'
        ], "Элемент 4")

        # Элемент 5 — ввод телефона
        try:
            phone_input=wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="phone-input"]')))
            phone_input.click()
            rnd="9"+"".join([str(random.randint(0,9)) for _ in range(9)])
            phone_input.send_keys(rnd)
            print(f"Введён номер: {rnd}")
            pause()
        except Exception as e:
            print("Поле ввода телефона не найдено:", e)

        # Элемент 6
        click_multi(driver, [
            '//*[@id="root"]/div/div/div[2]/div/div/div[2]/div[2]/div/div/div/div[2]/div[4]/div[1]/div/button/p'
        ], "Элемент 6")

        # Сохраняем ссылку и номер
        # Отдельный, более длинный таймаут: после Элемента 6 страница с номером
        # может грузиться дольше 15 с. Если элемент так и не появился — это ошибка
        # ДО цикла подтверждения, и окно останется открытым (см. finally) для разбора.
        try:
            phone_text_elem = WebDriverWait(driver, RESERVE_TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH,'//*[@id="root"]/div/div/main/div/div[1]/div/div/p[2]'))
            )
        except TimeoutException:
            print(f"Номер не появился за {RESERVE_TIMEOUT} с после Элемента 6 "
                  f"(текущий URL: {driver.current_url}). Окно оставлено открытым для разбора.")
            raise
        digits=re.sub(r"[^\d+]","",phone_text_elem.text.strip()); current_url=driver.current_url
        with open(log_file,"a",encoding="utf-8") as f:
            f.write(f"{current_url}|{digits}\n")
        print(f"Ссылка и номер сохранены: {current_url}|{digits}")

        # Элемент 7
        click_multi(driver, ['//*[@id="root"]/div/div/main/div/div[2]/div/div/button/p'], "Элемент 7")

        # Элемент 8 + ввод логина/пароля
        submitted = False
        while row_index < len(rows):
            click_multi(driver, ['//*[@id="root"]/div/div/main/div/div[1]/div/div[2]/div[1]/div/div'], "Элемент 8")

            num, pas = str(rows[row_index][0]), str(rows[row_index][1])
            phone_field = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="root"]/div/div/main/div/div/form/div[1]/div[1]/div/div/div[1]/input')))
            phone_field.clear(); phone_field.send_keys(num)
            pass_field = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="root"]/div/div/main/div/div/form/div[1]/div[2]/div/div/div[1]/input')))
            pass_field.clear(); pass_field.send_keys(pas)

            time.sleep(4)
            p_els = driver.find_elements(By.XPATH,'//*[@id="root"]/div/div/main/div/div/form/div[2]/button/p')
            submit_p = p_els[0] if p_els else None
            submit_btn = submit_p.find_element(By.XPATH, '..') if submit_p else None

            if submit_btn and is_ready(submit_btn):
                submit_btn.click()
                print(f"Клик по Элемент 9 (клиент: {num}|{pas})")
                submitted = True
                break
            else:
                print(f"Элемент 9 не кликабелен для {num}|{pas} — пробуем следующего клиента")
                row_index += 1

        if not submitted:
            print("Клиенты закончились — Элемент 9 так и не стал кликабельным")
            break

        # --- ожидание загрузки mobile-id-auth ---
        # ВАЖНО: таймаут здесь НЕ обрывает окно — цикл подтверждения запускается в любом случае.
        try:
            WebDriverWait(driver, AUTH_PAGE_TIMEOUT).until(
                EC.url_contains("mobile-id-auth")
            )
            print(f"Страница mobile-id-auth загружена: {driver.current_url}")
        except TimeoutException:
            print(f"Не дождался перехода на mobile-id-auth за {AUTH_PAGE_TIMEOUT} с "
                  f"(текущий URL: {driver.current_url}) — всё равно жду подтверждение")

        # --- цикл ожидания подтверждения (6 раз по 90 секунд) ---
        reached_confirmation = True
        window_ok = wait_confirmation(driver)

        if window_ok:
            with open(log_file,"a",encoding="utf-8") as f:
                f.write(f"{current_url}|{digits}|{num}|{pas}\n")
            print(f"Успешный лог сохранён: {current_url}|{digits}|{num}|{pas}")
            print("Окно оставлено открытым, перехожу к новому окну")
        else:
            with open(log_file,"a",encoding="utf-8") as f:
                f.write(f"FAIL|{current_url}|{digits}|{num}|{pas}\n")
            print(f"Окно неуспешное: {current_url}|{digits}|{num}|{pas}")

    except Exception as e:
        print(f"Ошибка ({type(e).__name__}): {e}")
        traceback.print_exc()
    finally:
        # Закрываем окно ТОЛЬКО если цикл подтверждения реально отработал и вернул неуспех.
        # Ошибка на любом шаге ДО цикла (например, номер не появился после Элемента 6)
        # оставляет окно открытым, чтобы можно было разобраться, что пошло не так.
        if reached_confirmation and not window_ok and CLOSE_FAILED_WINDOW:
            try:
                driver.quit()
                print("Неуспешное окно закрыто")
            except Exception:
                pass

    row_index += 1
